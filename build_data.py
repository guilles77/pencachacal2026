#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build de la Penca Chacal (app de SOLO visualizacion en Vercel).

Regenera la carpeta web/ a partir de:
  - index.html  (frontend canonico)
  - "Penca Los Chacales Mundial 2026.xlsx"  (usuarios, pronosticos, resultados, config)

Uso:
  python build_data.py
  vercel deploy --prod --yes

Cuando se juegan partidos: actualiza la hoja "Actual" (y/o "Predictions") del Excel,
volve a correr este script y redeploya. Los puntajes se recalculan solos.
"""
import io, json, os, sys
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(ROOT, "Penca Los Chacales Mundial 2026.xlsx")
WEB = os.path.join(ROOT, "web")

PUB_CONFIG_KEYS = {"PREDICTION_DEADLINE", "REVEAL_PREDICTIONS", "OFFICIAL_RESULTS_URL"}
PUB_CONFIG_PREFIX = ("POINT_",)
# Campos sensibles que NUNCA se publican (tokens, emails, claves, pin, api keys).


def build_data_js():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    players = []
    for row in wb["Players"].iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            players.append({"playerId": str(row[0]), "name": str(row[1])})
    preds = {}
    for row in wb["Predictions"].iter_rows(min_row=2, values_only=True):
        pid = row[0]
        if not pid:
            continue
        raw = row[2] or row[1]  # SubmittedJson, si no DraftJson
        if not raw:
            continue
        try:
            preds[str(pid)] = json.loads(raw)
        except Exception as e:
            print("WARN pred parse", pid, e)
    actual = {}
    for row in wb["Actual"].iter_rows(min_row=2, values_only=True):
        if row[0] == "actual" and row[1]:
            actual = json.loads(row[1])
            break
    config = {}
    for row in wb["Config"].iter_rows(min_row=2, values_only=True):
        k, v = row[0], row[1]
        if not k:
            continue
        k = str(k)
        if k in PUB_CONFIG_KEYS or k.startswith(PUB_CONFIG_PREFIX):
            if v is not None:
                if isinstance(v, float) and v == int(v):
                    v = int(v)
                config[k] = v
    bundle = {"config": config, "players": players, "predictions": preds, "actual": actual}
    js = "window.LC_STATIC_DATA = " + json.dumps(bundle, ensure_ascii=False) + ";\n"
    io.open(os.path.join(WEB, "data.js"), "w", encoding="utf-8").write(js)
    played = len([1 for k, v in actual.get("groupScores", {}).items()
                  if v and v.get("homeGoals") not in ("", None)])
    print("data.js  -> %d jugadores, %d pronosticos, %d grupos con resultado" % (len(players), len(preds), played))


def build_index_html():
    s = io.open(os.path.join(ROOT, "index.html"), encoding="utf-8").read()
    reps = {
        "<?!= initialParamsJson ?>": "{}",
        "<?!= serviceUrlJson ?>": '""',
        "<?!= pwaManifestUrlJson ?>": '""',
        "<?!= pwaServiceWorkerUrlJson ?>": '"/sw.js"',
        "<?!= pwaIconDataUrlJson ?>": '""',
    }
    for a, b in reps.items():
        s = s.replace(a, b)
    anchor = '<meta name="apple-mobile-web-app-title" content="Penca Chacal">'
    head = (anchor +
            '\n  <link rel="manifest" href="/manifest.webmanifest">'
            '\n  <link rel="apple-touch-icon" href="/icons/apple-touch-icon.png">'
            '\n  <link rel="icon" type="image/png" sizes="32x32" href="/icons/favicon-32.png">'
            '\n  <link rel="icon" type="image/png" sizes="16x16" href="/icons/favicon-16.png">'
            '\n  <meta name="lc-api-url" content="">'
            '\n  <script src="/data.js"></script>')
    if anchor not in s:
        print("ERROR: no encontre el anchor del head en index.html"); sys.exit(1)
    s = s.replace(anchor, head, 1)
    io.open(os.path.join(WEB, "index.html"), "w", encoding="utf-8").write(s)
    print("index.html -> web/index.html (%d KB)" % (len(s.encode("utf-8")) // 1024))


if __name__ == "__main__":
    os.makedirs(WEB, exist_ok=True)
    # El Excel tiene secretos y NO está versionado. Si está, regenera web/data.js;
    # si no, se mantiene el web/data.js ya commiteado y solo se regenera el HTML.
    if os.path.exists(XLSX):
        build_data_js()
    else:
        print("AVISO: no encuentro el Excel; mantengo web/data.js actual y solo regenero el HTML.")
    build_index_html()
    print("\nListo. Ahora deploya con:  vercel deploy --prod --yes")
